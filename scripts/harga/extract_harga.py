# -*- coding: utf-8 -*-
"""
PAAX - Extractor daftar harga bahan dan upah ke price book engine.

Kode masuk repo; hasil JSON price book dan audit ditulis ke PAAX_DATA_DIR
di luar repo (default: G:\\paax-data).
"""
from __future__ import annotations

import argparse
import ast
import csv
import json
import operator
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


SOURCE_DEFAULT = r"G:\AHSP\Daftar harga bahan dan upah.xlsx"
DATA_DEFAULT = r"G:\paax-data"
CATALOG_REL = r"harga-satuan\_resources_catalog.json"
AHSP_REL = r"ahsp\cipta-karya-2026.json"
REPO_ROOT = Path(__file__).resolve().parents[2]
OVERRIDES_DEFAULT = REPO_ROOT / "data" / "harga-satuan" / "semarang_overrides.json"
REVIEW_FIELDS = [
    "source_name",
    "source_unit",
    "source_category",
    "source_price",
    "candidate_1_code",
    "candidate_1_name",
    "candidate_2_code",
    "candidate_2_name",
    "candidate_3_code",
    "candidate_3_name",
    "candidate_4_code",
    "candidate_4_name",
    "candidate_5_code",
    "candidate_5_name",
    "chosen_code",
]


@dataclass(frozen=True)
class HargaRow:
    source_name: str
    unit: str
    price: float
    category: str
    row_number: int


_UNIT_MAP = {
    "bh": "buah",
    "buah": "buah",
    "btg": "batang",
    "batang": "batang",
    "lbr": "lembar",
    "lembar": "lembar",
    "ltr": "liter",
    "liter": "liter",
    "kg": "kg",
    "m": "m",
    "m'": "m",
    "m\u0142": "m3",
    "m\u2019": "m",
    "m2": "m2",
    "m\u00b2": "m2",
    "m3": "m3",
    "m\u00b3": "m3",
    "oh": "OH",
    "oj": "OJ",
}

_WORD_ALIASES = {
    "portland cement": "semen portland",
    "cement": "semen",
    "begisting": "bekisting",
    "sealent": "sealant",
    "stainles": "stainless",
    "plafond": "plafon",
}

_STOP_NAME_TOKENS = {"pc", "sp", "uk"}
_SAFE_EXTRA_TOKENS = {
    "biasa",
    "kerikil",
    "lokasi",
    "pekerjaan",
    "pelitur",
    "polos",
    "quarry",
    "ulir",
}

_OPS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.USub: operator.neg,
    ast.UAdd: operator.pos,
}


def _as_number(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        text = text.replace("Rp", "").replace("rp", "").replace(" ", "")
        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".")
        elif "," in text:
            text = text.replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return None
    return None


def _clean_number(value: float) -> float | int:
    return int(value) if abs(value - int(value)) < 1e-9 else round(value, 6)


def _eval_ast(node: ast.AST) -> float:
    if isinstance(node, ast.Expression):
        return _eval_ast(node.body)
    if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
        return float(node.value)
    if isinstance(node, ast.UnaryOp) and type(node.op) in _OPS:
        return float(_OPS[type(node.op)](_eval_ast(node.operand)))
    if isinstance(node, ast.BinOp) and type(node.op) in _OPS:
        return float(_OPS[type(node.op)](_eval_ast(node.left), _eval_ast(node.right)))
    raise ValueError("Formula harga hanya boleh berisi angka, referensi sel, dan operator + - * /.")


def resolve_formula_value(ws: Worksheet, cell_ref: str, _seen: set[str] | None = None) -> float | int:
    """Resolve nilai angka atau formula sederhana seperti =F12+10000."""
    coord = cell_ref.upper()
    seen = _seen or set()
    if coord in seen:
        raise ValueError(f"Referensi formula melingkar di {coord}")
    seen.add(coord)

    value = ws[coord].value
    number = _as_number(value)
    if number is not None:
        return _clean_number(number)

    if not isinstance(value, str) or not value.strip().startswith("="):
        raise ValueError(f"Sel {coord} tidak berisi angka/formula harga: {value!r}")

    expr = value.strip()[1:]

    def replace_cell(match: re.Match[str]) -> str:
        nested = resolve_formula_value(ws, match.group(0), seen)
        return str(nested)

    expanded = re.sub(r"\b[A-Z]{1,3}\d+\b", replace_cell, expr.upper())
    if re.search(r"[^0-9+\-*/(). ]", expanded):
        raise ValueError(f"Formula tidak didukung di {coord}: {value!r}")
    result = _eval_ast(ast.parse(expanded, mode="eval"))
    return _clean_number(result)


def normalize_unit(unit: str | None) -> str:
    if not unit:
        return ""
    raw = str(unit).strip().replace(" ", "")
    raw = raw.replace("\u2032", "'").replace("\u2018", "'").replace("\u2019", "'")
    lowered = raw.lower()
    return _UNIT_MAP.get(lowered, raw)


def normalize_name(text: str | None) -> str:
    if not text:
        return ""
    value = unicodedata.normalize("NFKD", str(text).lower())
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.replace("\u00f8", " diameter ").replace("\u03c6", " diameter ")
    value = value.replace("(", " ").replace(")", " ")
    value = re.sub(r"[^a-z0-9]+", " ", value)
    value = re.sub(r"\b\d+(?:\s*x\s*\d+)+(?:\s*x\s*\d+)?\b", " ", value)
    value = re.sub(r"\b\d+\b", " ", value)
    value = re.sub(r"\b(cm|mm|m|kg|bh|buah|btg|batang|liter|lembar|oh|oj)\b", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    for src, dst in _WORD_ALIASES.items():
        value = re.sub(rf"\b{re.escape(src)}\b", dst, value)
    for token in _STOP_NAME_TOKENS:
        value = re.sub(rf"\b{re.escape(token)}\b", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def _number_signature(text: str | None) -> set[str]:
    if not text:
        return set()
    value = unicodedata.normalize("NFKD", str(text).lower()).replace(",", ".")
    k_marks = {f"k{match}" for match in re.findall(r"\bk\s*[-=]?\s*(\d{2,4})\b", value)}
    numbers = set(re.findall(r"\d+(?:\.\d+)?", value))
    return k_marks | numbers


def _numbers_compatible(source_name: str, resource_name: str) -> bool:
    source_numbers = _number_signature(source_name)
    resource_numbers = _number_signature(resource_name)
    return not source_numbers or not resource_numbers or source_numbers == resource_numbers


def parse_harga_sheet(ws: Worksheet) -> List[HargaRow]:
    rows: List[HargaRow] = []
    category: str | None = None
    for row in range(1, ws.max_row + 1):
        name_cell = ws.cell(row=row, column=2).value
        name = str(name_cell).strip() if name_cell is not None else ""
        marker = normalize_name(name)
        if marker == "upah":
            category = "upah"
            continue
        if marker == "bahan":
            category = "bahan"
            continue
        if category is None or not name:
            continue
        if marker in {"no", "uraian pekerja", "satuan", "harga"}:
            continue
        unit = normalize_unit(ws.cell(row=row, column=5).value)
        try:
            price = resolve_formula_value(ws, f"F{row}")
        except ValueError:
            continue
        price_num = _as_number(price)
        if price_num is None or price_num <= 0:
            continue
        rows.append(HargaRow(
            source_name=name,
            unit=unit,
            price=_clean_number(price_num),
            category=category,
            row_number=row,
        ))
    return rows


def load_price_rows(source_xlsx: Path) -> List[HargaRow]:
    wb = load_workbook(source_xlsx, data_only=False)
    try:
        ws = wb["Lembar1"] if "Lembar1" in wb.sheetnames else wb.active
        return parse_harga_sheet(ws)
    finally:
        wb.close()


def load_catalog(catalog_path: Path) -> List[Dict[str, Any]]:
    raw = json.loads(catalog_path.read_text(encoding="utf-8"))
    return list(raw.get("resources", []))


def load_usage_counts(ahsp_path: Path | None) -> Dict[str, int]:
    if ahsp_path is None or not ahsp_path.exists():
        return {}
    raw = json.loads(ahsp_path.read_text(encoding="utf-8"))
    counts: Dict[str, int] = {}
    for item in raw.get("items", []):
        codes = {
            component.get("resource_code")
            for component in item.get("components", [])
            if component.get("resource_code")
        }
        for code in codes:
            counts[code] = counts.get(code, 0) + 1
    return counts


def load_overrides(overrides_path: Path | None) -> Dict[str, str]:
    if overrides_path is None or not overrides_path.exists():
        return {}
    raw = json.loads(overrides_path.read_text(encoding="utf-8"))
    if isinstance(raw, dict) and isinstance(raw.get("overrides"), list):
        entries = raw["overrides"]
    elif isinstance(raw, list):
        entries = raw
    elif isinstance(raw, dict):
        return {str(k).strip(): str(v).strip() for k, v in raw.items() if str(k).strip() and str(v).strip()}
    else:
        raise ValueError(f"Format override tidak didukung: {overrides_path}")

    overrides: Dict[str, str] = {}
    for entry in entries:
        if not isinstance(entry, dict):
            raise ValueError(f"Entry override tidak valid di {overrides_path}: {entry!r}")
        source_name = str(entry.get("source_name", "")).strip()
        code = str(entry.get("code", "")).strip()
        if source_name and code:
            overrides[source_name] = code
    return overrides


def _candidate_score(source: HargaRow, resource: Dict[str, Any]) -> tuple[int, float]:
    if source.category != resource.get("category"):
        return 0, 0.0
    source_unit = normalize_unit(source.unit)
    resource_unit = normalize_unit(resource.get("unit", ""))
    if source_unit and resource_unit and source_unit != resource_unit:
        return 0, 0.0
    if not _numbers_compatible(source.source_name, str(resource.get("name", ""))):
        return 0, 0.0

    s_name = normalize_name(source.source_name)
    r_name = normalize_name(resource.get("name", ""))
    if not s_name or not r_name:
        return 0, 0.0
    if s_name == r_name:
        return 3, 1.0

    s_tokens = set(s_name.split())
    r_tokens = set(r_name.split())
    if s_tokens and s_tokens <= r_tokens and r_tokens - s_tokens <= _SAFE_EXTRA_TOKENS:
        return 2, len(s_tokens) / len(r_tokens)
    if r_tokens and r_tokens <= s_tokens and s_tokens - r_tokens <= _SAFE_EXTRA_TOKENS:
        return 2, len(r_tokens) / len(s_tokens)

    overlap = len(s_tokens & r_tokens)
    union = len(s_tokens | r_tokens) or 1
    score = overlap / union
    return (1, score) if score >= 0.92 and overlap >= 3 else (0, score)


def _match_one(
    source: HargaRow,
    catalog: List[Dict[str, Any]],
    usage_counts: Dict[str, int] | None = None,
) -> tuple[str, List[Dict[str, Any]], float, str, str]:
    usage_counts = usage_counts or {}
    scored = []
    for resource in catalog:
        tier, score = _candidate_score(source, resource)
        if tier:
            scored.append((tier, score, resource))
    if not scored:
        return "unmatched", [], 0.0, "none", "tidak ada kandidat aman"
    scored.sort(key=lambda item: (item[0], item[1]), reverse=True)
    best_tier, best_score, best_resource = scored[0]
    tied = [
        r for tier, score, r in scored
        if tier == best_tier and abs(score - best_score) < 1e-9
    ]
    unique_codes = {r.get("code") for r in tied}
    if len(unique_codes) > 1:
        used_codes = {
            code: usage_counts.get(str(code), 0)
            for code in unique_codes
            if usage_counts.get(str(code), 0) > 0
        }
        if len(used_codes) == 1:
            selected_code, count = next(iter(used_codes.items()))
            selected = next(r for r in tied if r.get("code") == selected_code)
            return (
                "matched",
                [selected],
                best_score,
                "usage_tiebreak",
                f"dipilih karena dipakai {count} item AHSP",
            )
        if len(used_codes) > 1:
            max_count = max(used_codes.values())
            winners = [code for code, count in used_codes.items() if count == max_count]
            if len(winners) == 1:
                selected_code = winners[0]
                selected = next(r for r in tied if r.get("code") == selected_code)
                detail = ", ".join(f"{code}={count}" for code, count in sorted(used_codes.items()))
                return (
                    "matched",
                    [selected],
                    best_score,
                    "usage_tiebreak",
                    f"dipilih karena dipakai {max_count} item AHSP dan paling sering dipakai ({detail})",
                )
        return "ambiguous", tied, best_score, "ambiguous", "beberapa kandidat aman tanpa tie-break unik"
    return "matched", [best_resource], best_score, "auto", f"auto tier {best_tier} score {round(best_score, 4)}"


def match_price_rows(
    rows: Iterable[HargaRow],
    catalog: List[Dict[str, Any]],
    usage_counts: Dict[str, int] | None = None,
    overrides: Dict[str, str] | None = None,
) -> tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    matched: List[Dict[str, Any]] = []
    unmatched: List[Dict[str, Any]] = []
    ambiguous: List[Dict[str, Any]] = []
    overrides = overrides or {}
    catalog_by_code = {str(resource.get("code")): resource for resource in catalog}
    for source in rows:
        base = {
            "source_name": source.source_name,
            "source_unit": source.unit,
            "source_category": source.category,
            "source_price": source.price,
            "source_row": source.row_number,
            "normalized_name": normalize_name(source.source_name),
            "match_score": 0.0,
        }
        if source.source_name in overrides:
            code = overrides[source.source_name]
            resource = catalog_by_code.get(code)
            if resource is None:
                raise ValueError(
                    f"Override harga tidak valid: {source.source_name!r} -> {code!r} tidak ada di katalog"
                )
            if source.category != resource.get("category"):
                raise ValueError(
                    f"Override harga tidak valid: {source.source_name!r} -> {code!r} beda kategori"
                )
            matched.append({
                **base,
                "match_score": 1.0,
                "match_method": "manual_override",
                "match_reason": "manual override dari file pemetaan tervalidasi kode dan kategori",
                "code": resource["code"],
                "name": resource["name"],
                "category": resource["category"],
                "unit": resource["unit"],
                "price": source.price,
            })
            continue

        status, candidates, score, method, reason = _match_one(source, catalog, usage_counts)
        base["match_score"] = round(score, 4)
        if status == "matched":
            resource = candidates[0]
            matched.append({
                **base,
                "match_method": method,
                "match_reason": reason,
                "code": resource["code"],
                "name": resource["name"],
                "category": resource["category"],
                "unit": resource["unit"],
                "price": source.price,
            })
        elif status == "ambiguous":
            ambiguous.append({
                **base,
                "match_method": method,
                "match_reason": reason,
                "candidates": [
                    {
                        "code": r.get("code"),
                        "name": r.get("name"),
                        "category": r.get("category"),
                        "unit": r.get("unit"),
                    }
                    for r in candidates
                ],
            })
        else:
            unmatched.append({**base, "match_method": method, "match_reason": reason})
    return matched, unmatched, ambiguous


def _review_candidate_score(source: Dict[str, Any], resource: Dict[str, Any]) -> tuple[int, int, float, int]:
    category_score = 1 if source.get("source_category") == resource.get("category") else 0
    source_unit = normalize_unit(source.get("source_unit", ""))
    resource_unit = normalize_unit(resource.get("unit", ""))
    unit_score = 1 if source_unit and resource_unit and source_unit == resource_unit else 0
    s_tokens = set(normalize_name(source.get("source_name", "")).split())
    r_tokens = set(normalize_name(resource.get("name", "")).split())
    overlap = len(s_tokens & r_tokens)
    union = len(s_tokens | r_tokens) or 1
    similarity = overlap / union
    return category_score, unit_score, similarity, overlap


def build_review_rows(
    pending: Iterable[Dict[str, Any]],
    catalog: List[Dict[str, Any]],
    limit: int = 5,
) -> List[Dict[str, Any]]:
    review_rows: List[Dict[str, Any]] = []
    for source in pending:
        row = {
            "source_name": source.get("source_name", ""),
            "source_unit": source.get("source_unit", ""),
            "source_category": source.get("source_category", ""),
            "source_price": source.get("source_price", ""),
            "chosen_code": "",
        }
        scored = []
        for resource in catalog:
            score = _review_candidate_score(source, resource)
            if score[2] <= 0 and not (score[0] and score[1]):
                continue
            scored.append((score, resource))
        scored.sort(key=lambda item: item[0], reverse=True)
        for idx, (_, resource) in enumerate(scored[:limit], start=1):
            row[f"candidate_{idx}_code"] = resource.get("code", "")
            row[f"candidate_{idx}_name"] = resource.get("name", "")
        for idx in range(1, limit + 1):
            row.setdefault(f"candidate_{idx}_code", "")
            row.setdefault(f"candidate_{idx}_name", "")
        review_rows.append(row)
    return review_rows


def build_price_book(
    matched: Iterable[Dict[str, Any]],
    region: str,
    region_code: str,
    source_file: str,
    effective_date: str = "2026-06-28",
) -> Dict[str, Any]:
    by_code: Dict[str, Dict[str, Any]] = {}
    for row in matched:
        by_code.setdefault(row["code"], {
            "code": row["code"],
            "name": row["name"],
            "category": row["category"],
            "unit": row["unit"],
            "price": row["price"],
        })
    return {
        "region": region,
        "region_code": region_code,
        "currency": "IDR",
        "source": f"Daftar harga bahan dan upah: {source_file}",
        "effective_date": effective_date,
        "resources": list(by_code.values()),
    }


def build_audit(
    rows: List[HargaRow],
    matched: List[Dict[str, Any]],
    unmatched: List[Dict[str, Any]],
    ambiguous: List[Dict[str, Any]],
    catalog: List[Dict[str, Any]],
    region_code: str,
    source_file: str,
) -> Dict[str, Any]:
    matched_codes = {row["code"] for row in matched}
    manual_override_rows = sum(1 for row in matched if row.get("match_method") == "manual_override")
    usage_tiebreak_rows = sum(1 for row in matched if row.get("match_method") == "usage_tiebreak")
    return {
        "region_code": region_code,
        "source_file": source_file,
        "policy": (
            "Match otomatis memakai category+unit, nama ternormalisasi exact/alias, "
            "token-subset/Jaccard ketat, manual override tervalidasi, dan tie-break "
            "kandidat ambigu berdasarkan pemakaian kode di AHSP. Kandidat domain tetap masuk review manual."
        ),
        "stats": {
            "source_rows": len(rows),
            "matched_rows": len(matched),
            "unique_matched_resources": len(matched_codes),
            "manual_override_rows": manual_override_rows,
            "usage_tiebreak_rows": usage_tiebreak_rows,
            "unmatched_rows": len(unmatched),
            "ambiguous_rows": len(ambiguous),
            "review_rows": len(unmatched) + len(ambiguous),
            "catalog_resources": len(catalog),
            "catalog_coverage_pct": round(len(matched_codes) / len(catalog) * 100, 4) if catalog else 0,
        },
        "matched": matched,
        "unmatched": unmatched,
        "ambiguous": ambiguous,
    }


def write_outputs(
    price_book: Dict[str, Any],
    audit: Dict[str, Any],
    review_rows: List[Dict[str, Any]],
    out_dir: Path,
    region_code: str,
) -> tuple[Path, Path, Path]:
    harga_dir = out_dir / "harga-satuan"
    audit_dir = out_dir / "_audit"
    harga_dir.mkdir(parents=True, exist_ok=True)
    audit_dir.mkdir(parents=True, exist_ok=True)

    price_path = harga_dir / f"{region_code}.json"
    audit_path = audit_dir / f"harga_{region_code}.json"
    review_path = audit_dir / f"harga_{region_code}_review.csv"
    price_path.write_text(json.dumps(price_book, ensure_ascii=False, indent=2), encoding="utf-8")
    audit_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    with review_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=REVIEW_FIELDS)
        writer.writeheader()
        writer.writerows(review_rows)
    return price_path, audit_path, review_path


def run_extract(
    source_xlsx: Path,
    catalog_path: Path,
    out_dir: Path,
    region: str,
    region_code: str,
    effective_date: str,
    ahsp_path: Path | None = None,
    overrides_path: Path | None = None,
) -> tuple[Path, Path, Path, Dict[str, Any]]:
    rows = load_price_rows(source_xlsx)
    catalog = load_catalog(catalog_path)
    usage_counts = load_usage_counts(ahsp_path)
    overrides = load_overrides(overrides_path)
    matched, unmatched, ambiguous = match_price_rows(rows, catalog, usage_counts=usage_counts, overrides=overrides)
    price_book = build_price_book(matched, region, region_code, str(source_xlsx), effective_date)
    audit = build_audit(rows, matched, unmatched, ambiguous, catalog, region_code, str(source_xlsx))
    audit["ahsp_usage_path"] = str(ahsp_path) if ahsp_path else ""
    audit["overrides_path"] = str(overrides_path) if overrides_path else ""
    pending = [*unmatched, *ambiguous]
    review_rows = build_review_rows(pending, catalog)
    price_path, audit_path, review_path = write_outputs(price_book, audit, review_rows, out_dir, region_code)
    audit["review_path"] = str(review_path)
    audit_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    return price_path, audit_path, review_path, audit


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default=SOURCE_DEFAULT)
    ap.add_argument("--catalog", default=str(Path(DATA_DEFAULT) / CATALOG_REL))
    ap.add_argument("--out", default=DATA_DEFAULT)
    ap.add_argument("--ahsp", default=str(Path(DATA_DEFAULT) / AHSP_REL))
    ap.add_argument("--overrides", default=str(OVERRIDES_DEFAULT))
    ap.add_argument("--region", default="Semarang")
    ap.add_argument("--region-code", default="semarang")
    ap.add_argument("--effective-date", default="2026-06-28")
    args = ap.parse_args()

    price_path, audit_path, review_path, audit = run_extract(
        source_xlsx=Path(args.src),
        catalog_path=Path(args.catalog),
        out_dir=Path(args.out),
        region=args.region,
        region_code=args.region_code,
        effective_date=args.effective_date,
        ahsp_path=Path(args.ahsp) if args.ahsp else None,
        overrides_path=Path(args.overrides) if args.overrides else None,
    )
    stats = audit["stats"]
    print("=== RINGKASAN HARGA ===")
    print(f"Source rows       : {stats['source_rows']}")
    print(f"Matched rows      : {stats['matched_rows']}")
    print(f"Unique resources  : {stats['unique_matched_resources']}")
    print(f"Manual overrides  : {stats['manual_override_rows']}")
    print(f"Usage tie-break   : {stats['usage_tiebreak_rows']}")
    print(f"Unmatched rows    : {stats['unmatched_rows']}")
    print(f"Ambiguous rows    : {stats['ambiguous_rows']}")
    print(f"Price book        : {price_path}")
    print(f"Audit             : {audit_path}")
    print(f"Review worksheet  : {review_path}")


if __name__ == "__main__":
    main()
