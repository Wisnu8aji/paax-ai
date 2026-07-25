"""Auditable backup worksheet export/import model for civil work items."""
from __future__ import annotations
from io import BytesIO
from typing import Iterable, Mapping
from openpyxl import Workbook, load_workbook

HEADERS=("Item pekerjaan","Lokasi/Lantai","Jenis","Satuan","Ukuran","Jumlah","Formula","Hasil","Status","Sumber")

def export_backup_workbook(items: Iterable[Mapping[str, object]]) -> bytes:
    wb=Workbook(); ws=wb.active; ws.title="Perhitungan Backup"; ws.append(HEADERS)
    for item in items:
        ws.append([item.get("display_name"),item.get("location"),item.get("category"),item.get("unit"),item.get("dimensions_display"),
                   item.get("count"),item.get("formula"),item.get("result_display"),item.get("status"),
                   "; ".join(f"Hal. {ref.get('page')} — {ref.get('role')}" for ref in item.get("source_refs",[]) if isinstance(ref, Mapping))])
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    stream=BytesIO(); wb.save(stream); return stream.getvalue()

def inspect_backup_workbook(payload: bytes) -> dict:
    wb=load_workbook(BytesIO(payload), read_only=True, data_only=False)
    ws=wb["Perhitungan Backup"]
    headers=tuple(cell.value for cell in next(ws.iter_rows(min_row=1,max_row=1)))
    if headers != HEADERS: raise ValueError("unsupported backup worksheet schema")
    return {"sheet":"Perhitungan Backup","row_count":max(0,ws.max_row-1),"headers":headers}
