from __future__ import annotations

import argparse
import io
import os
import json
import struct
import sys
import urllib.error
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DB = "http://127.0.0.1:8001"
CORE = "http://127.0.0.1:8081"
DOC = "http://127.0.0.1:8083"


def request_json(url: str, *, method: str = "GET", payload: Any = None, key: str, actor: str) -> Any:
    data = None if payload is None else json.dumps(payload).encode("utf-8")
    headers = {"X-Internal-Key": key, "X-User-Id": actor, "Content-Type": "application/json"}
    request = urllib.request.Request(url, data=data, headers=headers, method=method)
    with urllib.request.urlopen(request, timeout=60) as response:
        return json.load(response)


def request_bytes(url: str, *, key: str, actor: str) -> tuple[bytes, dict[str, str]]:
    request = urllib.request.Request(url, headers={"X-Internal-Key": key, "X-User-Id": actor})
    with urllib.request.urlopen(request, timeout=60) as response:
        return response.read(), dict(response.headers.items())


def png_size(payload: bytes) -> tuple[int, int]:
    if payload[:8] != b"\x89PNG\r\n\x1a\n" or payload[12:16] != b"IHDR":
        raise ValueError("not a PNG")
    return struct.unpack(">II", payload[16:24])


def fact(identifier: str, measurement_type: str, value: Any, unit: str, formula_input: str) -> dict[str, Any]:
    return {
        "measurement_id": identifier,
        "project_id": "PLHUT-SURAKARTA",
        "snapshot_id": "portable-phase30",
        "measurement_type": measurement_type,
        "value": value,
        "unit": unit,
        "source_method": "verified_instances" if measurement_type == "count" else "written_dimension",
        "formula_inputs": [formula_input],
        "verification_status": "engine_verified",
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Run live Phase 01-30 acceptance checks against PAAX services.")
    parser.add_argument("--internal-key", default=os.getenv("INTERNAL_SERVICE_KEY", "live-test-key"))
    parser.add_argument("--actor", default="paax-web")
    parser.add_argument("--artifacts-dir", type=Path, default=Path("data/portable/verification"))
    args = parser.parse_args()
    args.artifacts_dir.mkdir(parents=True, exist_ok=True)

    results: list[dict[str, Any]] = []

    def check(name: str, condition: bool, detail: Any) -> None:
        results.append({"name": name, "status": "PASS" if condition else "FAIL", "detail": detail})
        if not condition:
            raise AssertionError(f"{name}: {detail}")

    try:
        health = request_json(f"{DB}/health", key=args.internal_key, actor=args.actor)
        check("db_health", health.get("status") == "ok", health)
        core_health = request_json(f"{CORE}/health", key=args.internal_key, actor=args.actor)
        check("core_health", core_health.get("status") in {"ok", "healthy"}, core_health)
        doc_health = request_json(f"{DOC}/health", key=args.internal_key, actor=args.actor)
        check("document_intelligence_health", doc_health.get("status") in {"ok", "healthy"}, doc_health)

        projects = request_json(f"{DB}/projects", key=args.internal_key, actor=args.actor)
        plhut = next((project for project in projects if project.get("id") == "PLHUT-SURAKARTA"), None)
        check("plhut_registered", plhut is not None, plhut or projects)
        check("plhut_owner_binding", plhut.get("owner_id") == args.actor, plhut)

        manifest = request_json(f"{DB}/projects/PLHUT-SURAKARTA/source-document/manifest", key=args.internal_key, actor=args.actor)
        check("source_manifest_88_pages", manifest.get("page_count") == 88, manifest)
        check("source_manifest_checksum", manifest.get("sha256") == "bf582e74951312cc6ccd305c2d48772ca27e7ffdf5b0fb1a0ef7104c19e9eb68", manifest)

        sheets = request_json(f"{DB}/projects/PLHUT-SURAKARTA/dem/sheets", key=args.internal_key, actor=args.actor)
        check("drawing_intelligence_88_sheets", len(sheets) == 88, len(sheets))
        page_png, image_headers = request_bytes(
            f"{DB}/projects/PLHUT-SURAKARTA/source-document/pages/42/image?width=1400",
            key=args.internal_key,
            actor=args.actor,
        )
        width, height = png_size(page_png)
        (args.artifacts_dir / "PLHUT-HALAMAN-43-DENAH-KOLOM-L2.png").write_bytes(page_png)
        check("real_pdf_page_render", width >= 1000 and height >= 600, {"width": width, "height": height, "headers": image_headers})

        civil = request_json(f"{DB}/projects/PLHUT-SURAKARTA/project-graph/civil-work-items", key=args.internal_key, actor=args.actor)
        check("civil_work_item_projection", civil.get("summary") == {"total": 8, "ready": 7, "needs_review": 1, "by_location": {"Lantai 1": 4, "Lantai 2": 4}}, civil.get("summary"))
        k2 = next(item for item in civil["items"] if item["id"] == "work-column-K2-L2")
        check("k2_l2_user_projection", all((k2["count"] == 4, k2["dimensions_display"] == "0,250 × 0,600 × 3,900 m", k2["result_display"] == "2,340 m³")), k2)

        context = request_json(
            f"{DB}/projects/PLHUT-SURAKARTA/project-graph/engineering-context",
            method="POST",
            payload={"query": "Berapa volume kolom K2 lantai 2?"},
            key=args.internal_key,
            actor=args.actor,
        )
        check("command_room_project_binding", context.get("project_binding", {}).get("project_id") == "PLHUT-SURAKARTA", context.get("project_binding"))
        check("command_room_quantity_authority", context.get("quantity_authority") == "core_engine", context.get("quantity_authority"))
        check("command_room_claim_evidence", len(context.get("citations", [])) == 3 and context["facts"][0]["result"] == "2,340 m³", context)

        unknown = request_json(
            f"{DB}/projects/PLHUT-SURAKARTA/project-graph/engineering-context",
            method="POST",
            payload={"query": "Berapa volume kolom K9 lantai 2?"},
            key=args.internal_key,
            actor=args.actor,
        )
        check("command_room_abstention", unknown.get("matched_item_count") == 0 and unknown.get("quantity_authority") == "none" and bool(unknown.get("forbidden_claims")), unknown)

        calculation_payload = {
            "project_id": "PLHUT-SURAKARTA",
            "snapshot_id": "portable-phase30",
            "measurement_fact_ids": ["K2-W", "K2-D", "K2-H", "K2-C"],
            "calculation_type": "concrete_column_total_volume",
            "inputs": [
                fact("K2-W", "length", 250, "mm", "width"),
                fact("K2-D", "length", 600, "mm", "depth"),
                fact("K2-H", "length", 3900, "mm", "height"),
                fact("K2-C", "count", 4, "unit", "count"),
            ],
            "requested_by": args.actor,
        }
        calculated = request_json(f"{CORE}/calculations", method="POST", payload=calculation_payload, key=args.internal_key, actor=args.actor)
        check("core_engine_exact_calculation", calculated.get("status") == "complete" and calculated.get("result") == 2.34 and calculated.get("unit") == "m3", calculated)

        xlsx, _ = request_bytes(f"{DB}/projects/PLHUT-SURAKARTA/project-graph/civil-work-items/export.xlsx", key=args.internal_key, actor=args.actor)
        xlsx_path = args.artifacts_dir / "PAAX-PLHUT-PERHITUNGAN-BACKUP.xlsx"
        xlsx_path.write_bytes(xlsx)
        workbook = load_workbook(io.BytesIO(xlsx), data_only=False)
        sheet = workbook["Perhitungan Backup"]
        headers = [cell.value for cell in sheet[1]]
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        k2_row = next(row for row in rows if row[0] == "Kolom Beton Bertulang K2" and row[1] == "Lantai 2")
        check("calculation_backup_excel", headers == ["Item pekerjaan", "Lokasi/Lantai", "Jenis", "Satuan", "Ukuran", "Jumlah", "Formula", "Hasil", "Status", "Sumber"] and abs(float(k2_row[7]) - 2.34) < 1e-9 and sheet.cell(row=rows.index(k2_row)+2, column=8).number_format == '0.000 "m³"', {"headers": headers, "k2_row": k2_row})

    except (AssertionError, urllib.error.URLError, KeyError, StopIteration, ValueError) as exc:
        results.append({"name": "runtime_exception", "status": "FAIL", "detail": str(exc)})

    summary = {
        "schema_version": "paax.phase30-runtime-verification.v1",
        "status": "PASS" if all(item["status"] == "PASS" for item in results) else "FAIL",
        "passed": sum(item["status"] == "PASS" for item in results),
        "failed": sum(item["status"] == "FAIL" for item in results),
        "results": results,
    }
    report_path = args.artifacts_dir / "PHASE30_RUNTIME_VERIFICATION.json"
    report_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0 if summary["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
