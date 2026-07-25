from __future__ import annotations

import io
import json
import sys
import urllib.request
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

BASE = sys.argv[1].rstrip('/') if len(sys.argv) > 1 else 'http://127.0.0.1:8099'
checks: list[dict[str, object]] = []


def get(path: str) -> tuple[int, str, bytes]:
    req = urllib.request.Request(BASE + path)
    with urllib.request.urlopen(req, timeout=30) as r:
        return r.status, r.headers.get_content_type(), r.read()


def post(path: str, payload: dict) -> tuple[int, str, bytes]:
    req = urllib.request.Request(BASE + path, data=json.dumps(payload).encode(), method='POST', headers={'content-type':'application/json'})
    with urllib.request.urlopen(req, timeout=30) as r:
        return r.status, r.headers.get_content_type(), r.read()


def check(name: str, ok: bool, detail: str) -> None:
    checks.append({'name': name, 'ok': bool(ok), 'detail': detail})

status, ctype, body = get('/')
html = body.decode('utf-8')
check('workspace_html', status == 200 and 'PLHUT Surakarta — Aktif' in html, f'{status} {ctype}')
check('professional_quantity_headers', all(h in html for h in ['Item pekerjaan','Lokasi/Lantai','Satuan','Ukuran','Jumlah','Formula','Volume/Hasil','Sumber']), 'user-facing columns')
check('command_room_visible', 'Verified Engineering Context' in html, 'Command Room acceptance surface')

_, _, body = get('/api/manifest')
manifest = json.loads(body)
check('manifest_project', manifest.get('project_id') == 'PLHUT-SURAKARTA', str(manifest.get('project_id')))
check('manifest_88_pages', manifest.get('page_count') == 88, str(manifest.get('page_count')))

_, _, body = get('/api/items')
items = json.loads(body)
check('quantity_summary', items.get('summary') == {'total':8,'ready':7,'needs_review':1,'by_location':{'Lantai 1':4,'Lantai 2':4}}, json.dumps(items.get('summary'), ensure_ascii=False))
k2 = next((x for x in items.get('items',[]) if x.get('technical_code')=='K2' and x.get('location')=='Lantai 2'), None)
check('k2_projection', bool(k2 and k2.get('result') == 2.34 and k2.get('source_pages') == [43,50,54]), json.dumps(k2, ensure_ascii=False)[:300])

_, ctype, body = get('/api/page/42/image?width=1800')
image = Image.open(io.BytesIO(body))
check('real_pdf_image', ctype == 'image/png' and image.width >= 1400 and image.height > 900, f'{ctype} {image.size}')

_, _, body = post('/api/command', {'query':'Berapa volume Kolom K2 Lantai 2?'})
answer = json.loads(body)
check('command_project_binding', answer.get('project_binding',{}).get('project_id') == 'PLHUT-SURAKARTA', json.dumps(answer.get('project_binding')))
check('command_verified_answer', '2,340 m³' in answer.get('answer','') and answer.get('quantity_authority') == 'core_engine', answer.get('answer',''))
check('command_evidence', all(str(page) in answer.get('evidence_summary','') for page in [43,50,54]), answer.get('evidence_summary',''))

_, ctype, body = get('/api/export.xlsx')
wb = load_workbook(io.BytesIO(body), data_only=False)
ws = wb['Perhitungan Backup']
headers = [cell.value for cell in ws[1]]
check('excel_headers', headers == ['Item pekerjaan','Lokasi/Lantai','Jenis','Satuan','Ukuran','Jumlah','Formula','Hasil','Status','Sumber'], str(headers))
check('excel_rows', ws.max_row == 9, f'{ws.max_row} rows')
check('excel_k2', any(row[0].value == 'Kolom Beton Bertulang K2' and row[1].value == 'Lantai 2' and row[7].value == 2.34 for row in ws.iter_rows(min_row=2)), 'K2 L2 = 2.34')

failed = [x for x in checks if not x['ok']]
report = {'schema_version':'paax.phase30.acceptance-ui.v1','status':'FAIL' if failed else 'PASS','passed':len(checks)-len(failed),'failed':len(failed),'checks':checks}
print(json.dumps(report, ensure_ascii=False, indent=2))
raise SystemExit(1 if failed else 0)
