# -*- coding: utf-8 -*-
"""
Generator price book Surakarta NYATA dari RAB manual ALFA.xlsx (owner-authorized
2026-07-03: "harga DKH/HSP/HARGA BAHAN itu harga asli Surakarta, pakai saja").

Pemakaian:
    python _generate_surakarta_from_alfa.py "<path ALFA.xlsx>"

Sumber HSD (Harga Satuan Dasar) Surakarta 2024:
 - sheet HARGA BAHAN = daftar harga bahan+upah kanonik (AUTORITATIF)
 - + resource yang dipakai 32 AHS tapi belum ada di HARGA BAHAN (harga inline)
Resource dikunci by NAMA (kode ALFA tak andal — lihat fixture README). Untuk 5
resource yang harganya beda antara HARGA BAHAN vs analisa (inkonsistensi internal
ALFA), HARGA BAHAN menang (daftar harga = otoritas; RULE-HRG-02) dan dicatat.

Ini price book UMUM regional Surakarta (dipakai proyek Surakarta mana pun) — sah
sbg grounding sistem per prinsip §0.1 roadmap (harga = pengecualian; yang dilarang
jadi template = koefisien/answer-key PLHUT, itu tetap di tests/fixtures/).

CATATAN (ditemukan 2026-07-03): ada file SERUPA di luar repo,
`G:\\paax-data\\harga-satuan\\surakarta.json` (109 resource, sumber sama — HARGA
BAHAN ALFA — dibangun sesi sebelumnya), dipakai HANYA bila env `PAAX_DATA_DIR`
diset (lihat services/core-engine/app/rab/loader.py). Kode SKA.* di file itu
ASSIGNED BEDA (mis. SKA.L.001="Pekerja" di sana, ="1/3 GALIAN" di sini) — TIDAK
saling menimpa (tak pernah dimuat bersamaan; loader pilih salah satu tergantung
env), tapi berpotensi bingung bila dua developer memakai env berbeda. Belum
direkonsiliasi (file eksternal di luar jangkauan git repo ini) — item tindak
lanjut, bukan bug aktif. Lihat docs/plans/PAAX_FASE0B_GAP_HARGA_2026-07-03.md.
"""
from __future__ import annotations
import json, re, sys, unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("Butuh openpyxl")

HERE = Path(__file__).resolve().parent
OUT = HERE / "surakarta.json"


def norm(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().lower()
    return re.sub(r"[^a-z0-9]+", " ", s).strip()


CAT = {"upah": "upah", "tenaga": "upah", "bahan": "bahan", "peralatan": "alat", "alat": "alat"}
CODE_PREFIX = {"upah": "SKA.L", "bahan": "SKA.M", "alat": "SKA.E"}


def build(src):
    wb = openpyxl.load_workbook(src, data_only=True)

    # ---- HARGA BAHAN (autoritatif) ----
    ws = wb["HARGA BAHAN"]
    def s(r, c):
        v = ws.cell(row=r, column=c).value
        return "" if v is None else str(v).strip()
    def n(r, c):
        v = ws.cell(row=r, column=c).value
        return float(v) if isinstance(v, (int, float)) else None

    resources = {}   # norm_name -> dict
    cur = None
    for r in range(11, ws.max_row + 1):
        lab, price, unit = s(r, 2), n(r, 6), s(r, 5)
        ll = norm(lab)
        if ll == "upah": cur = "upah"; continue
        if ll == "bahan": cur = "bahan"; continue
        if ll in ("alat", "peralatan"): cur = "alat"; continue
        if lab and price is not None and price > 0:
            cat = cur or "bahan"
            resources[norm(lab)] = {"name": lab, "category": cat, "unit": unit or "-",
                                    "price": price, "source": "HARGA BAHAN"}

    # ---- resource dari AHS yg belum ada (harga inline) + catat konflik ----
    wsa = wb["AHS"]
    def sa(r, c):
        v = wsa.cell(row=r, column=c).value
        return "" if v is None else str(v).strip()
    def na(r, c):
        v = wsa.cell(row=r, column=c).value
        return float(v) if isinstance(v, (int, float)) else None
    def cat_of(l3):
        m = re.sub(r"\s+", "", l3.lower()); return next((v for k, v in CAT.items() if k in m), None)

    maxr = wsa.max_row
    starts = [r for r in range(1, maxr + 1) if sa(r, 2) == "Jenis Pekerjaan"] + [maxr + 1]
    conflicts = []
    for bi in range(len(starts) - 1):
        r0, r1 = starts[bi], starts[bi + 1]; cat = None
        for r in range(r0, r1):
            l2, l3 = sa(r, 2), sa(r, 3)
            if l2 in ("A.", "B.", "C."): cat = cat_of(l3); continue
            if l2 == "D": cat = None; continue
            if l2 == "F": break
            coef, nm, price = na(r, 9), sa(r, 4), na(r, 10)
            if cat and coef is not None and nm and price is not None:
                key = norm(nm)
                if key in resources:
                    hbp = resources[key]["price"]
                    if abs(price - hbp) > max(1.0, 0.005 * hbp) and key not in {norm(c["name"]) for c in conflicts}:
                        conflicts.append({"name": nm, "harga_bahan": hbp, "analisa": price})
                else:
                    resources[key] = {"name": nm, "category": cat, "unit": sa(r, 8) or "-",
                                      "price": price, "source": "AHS inline"}

    # ---- assign kode Surakarta stabil per kategori ----
    counters = {"upah": 0, "bahan": 0, "alat": 0}
    out = []
    for key in sorted(resources):
        r = resources[key]
        counters[r["category"]] += 1
        code = f"{CODE_PREFIX[r['category']]}.{counters[r['category']]:03d}"
        out.append({"code": code, "name": r["name"], "category": r["category"],
                    "unit": r["unit"], "price": r["price"], "norm_name": key})

    doc = {
        "region": "Kota Surakarta",
        "region_code": "surakarta",
        "currency": "IDR",
        "source": "RAB manual Gedung PLHUT Surakarta TA 2024 (ALFA.xlsx: HARGA BAHAN + AHS) — harga nyata Surakarta 2024, owner-authorized 2026-07-03",
        "effective_date": "2024",
        "note": ("Price book HSD Surakarta UMUM (proyek Surakarta mana pun). Prinsip roadmap Sec 0.1: "
                 "harga = grounding regional sah; koefisien/answer-key PLHUT tetap di tests/fixtures/. "
                 f"{len(conflicts)} resource inkonsisten di ALFA (HARGA BAHAN vs analisa) -> HARGA BAHAN menang (RULE-HRG-02). "
                 "CATATAN: ada file serupa (109 resource) di G:\\paax-data\\harga-satuan\\surakarta.json "
                 "(dipakai bila env PAAX_DATA_DIR diset) dengan penomoran kode SKA.* BERBEDA -- belum "
                 "direkonsiliasi (di luar repo git). Lihat docs/plans/PAAX_FASE0B_GAP_HARGA_2026-07-03.md."),
        "alfa_price_conflicts": conflicts,
        "resources": out,
    }
    return doc


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    doc = build(sys.argv[1])
    OUT.write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK — {len(doc['resources'])} resource Surakarta ditulis ke {OUT} "
          f"({len(doc['alfa_price_conflicts'])} konflik dicatat)")
