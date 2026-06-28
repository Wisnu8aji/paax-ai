from pathlib import Path
import sys

from openpyxl import Workbook

REPO_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(REPO_ROOT))

import scripts.harga.extract_harga as harga_extract  # noqa: E402
from scripts.harga.extract_harga import (  # noqa: E402
    HargaRow,
    build_price_book,
    match_price_rows,
    normalize_name,
    parse_harga_sheet,
    resolve_formula_value,
)


def test_resolve_formula_value_handles_relative_price_formulas():
    wb = Workbook()
    ws = wb.active
    ws["F12"] = 102000
    ws["F13"] = "=F12+10000"
    ws["F18"] = "=F12+5000"
    ws["F19"] = "=F18+2000"

    assert resolve_formula_value(ws, "F13") == 112000
    assert resolve_formula_value(ws, "F19") == 109000


def test_normalize_name_and_match_catalog_exact_or_alias():
    source = HargaRow(
        source_name="Portland cement",
        unit="kg",
        price=1200,
        category="bahan",
        row_number=52,
    )
    catalog = [
        {"code": "M.GEN.0004", "name": "Semen Portland (PC)", "category": "bahan", "unit": "kg", "price": 0},
        {"code": "M.GEN.9999", "name": "Semen warna", "category": "bahan", "unit": "kg", "price": 0},
    ]

    assert normalize_name(" Semen Portland (PC) ") == "semen portland"
    matched, unmatched, ambiguous = match_price_rows([source], catalog)

    assert unmatched == []
    assert ambiguous == []
    assert matched[0]["code"] == "M.GEN.0004"
    assert matched[0]["price"] == 1200
    assert matched[0]["name"] == "Semen Portland (PC)"


def test_parse_harga_sheet_and_build_price_book_dedupes_by_resource_code():
    wb = Workbook()
    ws = wb.active
    ws.title = "Lembar1"
    ws["B11"] = "UPAH"
    ws["B12"] = "Pekerja "
    ws["E12"] = "OH"
    ws["F12"] = 102000
    ws["B13"] = "Mandor "
    ws["E13"] = "OH"
    ws["F13"] = "=F12+10000"
    ws["B27"] = "BAHAN"
    ws["B28"] = "Bata merah 5 x 11 x 22 cm"
    ws["E28"] = "buah"
    ws["F28"] = 500

    rows = parse_harga_sheet(ws)
    catalog = [
        {"code": "L.01", "name": "Pekerja", "category": "upah", "unit": "OH", "price": 0},
        {"code": "L.04", "name": "Mandor", "category": "upah", "unit": "OH", "price": 0},
        {"code": "M.BATA", "name": "Bata merah 5 x 11 x 22 cm", "category": "bahan", "unit": "buah", "price": 0},
    ]
    matched, unmatched, ambiguous = match_price_rows(rows, catalog)
    price_book = build_price_book(
        matched,
        region="Semarang",
        region_code="semarang",
        source_file="fixture.xlsx",
    )

    assert unmatched == []
    assert ambiguous == []
    assert [r["code"] for r in price_book["resources"]] == ["L.01", "L.04", "M.BATA"]
    assert [r["price"] for r in price_book["resources"]] == [102000, 112000, 500]


def test_matcher_rejects_broad_subset_and_conflicting_dimensions():
    rows = [
        HargaRow(
            source_name="Cat penutup (cat tembok exterior)",
            unit="kg",
            price=70000,
            category="bahan",
            row_number=96,
        ),
        HargaRow(
            source_name="Plywood 9 mm (120x240)",
            unit="lembar",
            price=70000,
            category="bahan",
            row_number=38,
        ),
    ]
    catalog = [
        {"code": "M.CAT.KAYU", "name": "Cat Penutup Kayu", "category": "bahan", "unit": "kg", "price": 0},
        {"code": "M.PLY.4", "name": "Plywood 4 mm, 120 x 240", "category": "bahan", "unit": "lembar", "price": 0},
    ]

    matched, unmatched, ambiguous = match_price_rows(rows, catalog)

    assert matched == []
    assert ambiguous == []
    assert [row["source_name"] for row in unmatched] == [
        "Cat penutup (cat tembok exterior)",
        "Plywood 9 mm (120x240)",
    ]


def test_usage_tiebreak_selects_candidate_used_by_ahsp(tmp_path):
    ahsp = {
        "items": [
            {"components": [{"resource_code": "M.SEMEN.RESMI"}]},
            {"components": [{"resource_code": "M.SEMEN.RESMI"}]},
            {"components": [{"resource_code": "M.SEMEN.JARANG"}]},
        ]
    }
    ahsp_path = tmp_path / "ahsp.json"
    ahsp_path.write_text(__import__("json").dumps(ahsp), encoding="utf-8")
    usage_counts = harga_extract.load_usage_counts(ahsp_path)
    source = HargaRow(
        source_name="Portland cement",
        unit="kg",
        price=1200,
        category="bahan",
        row_number=52,
    )
    catalog = [
        {"code": "M.SEMEN.PC", "name": "Semen Portland (PC)", "category": "bahan", "unit": "kg", "price": 0},
        {"code": "M.SEMEN.RESMI", "name": "Semen Portland", "category": "bahan", "unit": "kg", "price": 0},
        {"code": "M.SEMEN.JARANG", "name": "Semen Portland (SP)", "category": "bahan", "unit": "kg", "price": 0},
    ]

    matched, unmatched, ambiguous = match_price_rows([source], catalog, usage_counts=usage_counts)

    assert unmatched == []
    assert ambiguous == []
    assert matched[0]["code"] == "M.SEMEN.RESMI"
    assert matched[0]["match_method"] == "usage_tiebreak"
    assert "dipakai 2 item AHSP" in matched[0]["match_reason"]


def test_manual_override_has_priority_and_invalid_code_fails():
    source = HargaRow(
        source_name="Kloset jongkok porselen",
        unit="buah",
        price=350000,
        category="bahan",
        row_number=71,
    )
    catalog = [
        {"code": "M.KLOSET", "name": "Kloset Jongkok", "category": "bahan", "unit": "buah", "price": 0},
        {"code": "M.PORSELEN", "name": "Porselen 11x11", "category": "bahan", "unit": "buah", "price": 0},
    ]

    matched, unmatched, ambiguous = match_price_rows(
        [source],
        catalog,
        overrides={"Kloset jongkok porselen": "M.KLOSET"},
    )

    assert unmatched == []
    assert ambiguous == []
    assert matched[0]["code"] == "M.KLOSET"
    assert matched[0]["match_method"] == "manual_override"

    try:
        match_price_rows([source], catalog, overrides={"Kloset jongkok porselen": "M.TIDAK.ADA"})
    except ValueError as exc:
        assert "Override harga tidak valid" in str(exc)
    else:
        raise AssertionError("invalid override should fail clearly")


def test_build_review_rows_outputs_top_candidates_and_empty_chosen_code():
    pending = [{
        "source_name": "Paku sekrup",
        "source_unit": "kg",
        "source_category": "bahan",
        "source_price": 16000,
        "source_row": 85,
        "normalized_name": "paku sekrup",
        "match_score": 0,
    }]
    catalog = [
        {"code": "M.PAKU.SEKRUP", "name": "Paku Sekrup", "category": "bahan", "unit": "kg", "price": 0},
        {"code": "M.PAKU", "name": "Paku", "category": "bahan", "unit": "kg", "price": 0},
        {"code": "M.KAYU", "name": "Kayu", "category": "bahan", "unit": "m3", "price": 0},
    ]

    rows = harga_extract.build_review_rows(pending, catalog, limit=2)

    assert rows[0]["source_name"] == "Paku sekrup"
    assert rows[0]["candidate_1_code"] == "M.PAKU.SEKRUP"
    assert rows[0]["candidate_1_name"] == "Paku Sekrup"
    assert rows[0]["candidate_2_code"] == "M.PAKU"
    assert rows[0]["chosen_code"] == ""
