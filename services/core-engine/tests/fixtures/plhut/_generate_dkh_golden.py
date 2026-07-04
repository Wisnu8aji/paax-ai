# -*- coding: utf-8 -*-
"""
Regenerator FIXTURE UJI golden PLHUT — RAB total (sheet DKH ALFA.xlsx).

Pemakaian:
    python _generate_dkh_golden.py "<path ke rab gedung plhut surakarta ALFA.xlsx>"

Menulis `dkh_golden.json`: 224 baris item (uraian, satuan, volume, harga_satuan
ALFA, dan ahs_code bila harga-satuannya cocok salah satu dari 32 analisa di
ahs_golden.json). + grand_total (kunci jawaban RAB) + ppn_rate.

FAKTA terverifikasi (ALFA internal): tiap baris total = vol x hs x (1+11%);
Σ(vol x hs) x 1.11 == grand_total = Rp 1.860.078.607 (224/224 baris konsisten).

Hanya 79/224 baris punya rincian AHS (32 analisa); 145 sisanya direct/lump-sum
(SMKK, APD, dsb.) tanpa breakdown koefisien di file -> di test dimodelkan sbg
pseudo-AHSP harga langsung (overhead 0). Jadi test 0a-2 = ASSEMBLY engine
(Σ vol×HSP + PPN) vs RAB profesional nyata; layer HSP-dari-koefisien diuji
terpisah di test_plhut_hsp_golden (0a-1).

Prinsip §0.1: FIXTURE UJI, bukan data/template sistem.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("Butuh openpyxl: pip install openpyxl")

HERE = Path(__file__).resolve().parent
OUT = HERE / "dkh_golden.json"
AHS = HERE / "ahs_golden.json"
PPN_RATE = 0.11


def build(src_path: str) -> dict:
    ws = openpyxl.load_workbook(src_path, data_only=True)["DKH"]
    ahs = json.loads(AHS.read_text(encoding="utf-8"))
    ahs_by_hsp: dict[int, str] = {}
    for a in ahs["analyses"]:
        ahs_by_hsp.setdefault(round(a["expected_hsp"]), a["code"])

    def c(r, col):
        return ws.cell(row=r, column=col).value

    lines: list[dict] = []
    sum_pretax = 0.0
    sum_total = 0.0
    seq = 0
    for r in range(10, ws.max_row + 1):
        ur, sat, vol, hs, pjt, total = (c(r, 1), c(r, 2), c(r, 3), c(r, 4), c(r, 5), c(r, 7))
        if not (isinstance(vol, (int, float)) and isinstance(hs, (int, float)) and vol > 0 and hs > 0):
            continue
        seq += 1
        pre = vol * hs
        sum_pretax += pre
        sum_total += total if isinstance(total, (int, float)) else 0
        lines.append({
            "seq": seq,
            "uraian": str(ur).strip(),
            "satuan": (str(sat).strip() if sat else "-"),
            "volume": float(vol),
            "harga_satuan": float(hs),
            "ahs_code": ahs_by_hsp.get(round(hs)),  # None bila direct/lump-sum
            "total_alfa": float(total) if isinstance(total, (int, float)) else None,
        })

    grand = round(sum_total, 2)
    mapped = sum(1 for x in lines if x["ahs_code"])

    doc = {
        "_note": ("GOLDEN FIXTURE UJI RAB total dari DKH ALFA.xlsx (PLHUT Surakarta 2024). "
                  "grand_total = kunci jawaban. FIXTURE UJI bukan data/template sistem (roadmap Sec 0.1). "
                  "79/224 baris punya AHS; 145 direct (dimodelkan pseudo-AHSP harga langsung di test)."),
        "source_file": "rab gedung plhut surakarta ALFA.xlsx :: sheet DKH",
        "ppn_rate": PPN_RATE,
        "n_lines": len(lines),
        "n_mapped_ahs": mapped,
        "subtotal_pretax": round(sum_pretax, 2),
        "grand_total": grand,
        "lines": lines,
    }

    # verifikasi internal ALFA: Σpretax x 1.11 == grand_total
    assert abs(sum_pretax * (1 + PPN_RATE) - grand) <= max(1.0, 0.001 * grand), \
        f"ALFA internal inconsistent: {sum_pretax*1.11:.0f} vs {grand:.0f}"
    return doc


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    doc = build(sys.argv[1])
    OUT.write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK — {doc['n_lines']} baris ({doc['n_mapped_ahs']} ber-AHS), "
          f"grand_total Rp {doc['grand_total']:,.0f}, ditulis ke {OUT}")
