# -*- coding: utf-8 -*-
"""
Regenerator FIXTURE UJI golden PLHUT dari RAB manual ALFA.xlsx (sheet AHS).

Pemakaian:
    python _generate_ahs_golden.py "<path ke rab gedung plhut surakarta ALFA.xlsx>"

Menulis `ahs_golden.json` di folder ini SETELAH verifikasi mandiri 32/32
(engine formula (A+B+C)x(1+OP) == HSP final baris F ALFA). Bila verifikasi gagal,
file TIDAK ditulis (no-silent-fix, brain INV-TKG-03).

CATATAN: file sumber ALFA.xlsx TIDAK disertakan di repo (data proyek nyata).
Skrip ini ada untuk provenans/auditabilitas — membuktikan fixture = ekstraksi
setia, bukan angka karangan. Perlu `openpyxl`.

Prinsip §0.1 (roadmap): output adalah FIXTURE UJI, bukan data/template sistem.
Kode resource ALFA tidak andal (dipakai ulang lintas analisa untuk material
berbeda) -> resource dikunci LOKAL per-analisa; identitas asli = nama+harga.
"""
from __future__ import annotations

import json
import re
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("Butuh openpyxl: pip install openpyxl")

HERE = Path(__file__).resolve().parent
OUT = HERE / "ahs_golden.json"

CAT = {"upah": "upah", "tenaga": "upah", "bahan": "bahan", "peralatan": "alat", "alat": "alat"}


def _cat_of(label: str) -> str | None:
    # header AHS sering ditulis berspasi ("B a h a n") -> hapus spasi dulu
    n = re.sub(r"\s+", "", label.lower())
    return next((v for k, v in CAT.items() if k in n), None)


def _slug(name: str) -> str:
    n = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode()
    return (re.sub(r"[^a-zA-Z0-9]+", "-", n).strip("-").lower() or "res")[:28]


def build(src_path: str) -> dict:
    ws = openpyxl.load_workbook(src_path, data_only=True)["AHS"]

    def s(r, c):
        v = ws.cell(row=r, column=c).value
        return "" if v is None else str(v).strip()

    def num(r, c):
        v = ws.cell(row=r, column=c).value
        return float(v) if isinstance(v, (int, float)) else None

    maxr = ws.max_row
    starts = [r for r in range(1, maxr + 1) if s(r, 2) == "Jenis Pekerjaan"] + [maxr + 1]

    resources: dict[str, dict] = {}
    analyses: list[dict] = []

    for bi in range(len(starts) - 1):
        r0, r1 = starts[bi], starts[bi + 1]
        acode = f"PLHUT-AHS-{bi + 1:02d}"
        name = s(r0, 6)
        unit = ""
        cat = None
        op = 0.0
        F = None
        comps: list[dict] = []
        ci = 0
        for r in range(r0, r1):
            l2 = s(r, 2)
            l3 = s(r, 3)
            if l2 == "Satuan Pekerjaan":
                unit = s(r, 6)
                continue
            if l2 in ("A.", "B.", "C."):
                cat = _cat_of(l3)
                continue
            if l2 == "D":
                cat = None
                continue
            if l2 == "E":
                op = num(r, 9) or 0.0
                continue
            if l2 == "F":
                F = num(r, 11)
                break
            coef = num(r, 9)
            nm = s(r, 4)
            price = num(r, 10)
            if cat and coef is not None and nm and price is not None:
                ci += 1
                key = f"{acode}#R{ci:02d}"
                comps.append({"resource_code": key, "category": cat, "coefficient": coef})
                resources[key] = {
                    "code": key, "name": nm, "category": cat,
                    "unit": s(r, 8) or "-", "price": price, "alfa_code": s(r, 7) or None,
                }
        if name and comps and F is not None:
            analyses.append({
                "code": acode, "name": name, "unit": unit or "-",
                "overhead_profit": round(op, 4), "components": comps,
                "expected_hsp": round(F, 2),
            })

    # verifikasi mandiri (harus 32/32)
    pb = {k: v["price"] for k, v in resources.items()}
    bad = []
    for a in analyses:
        base = sum(c["coefficient"] * pb[c["resource_code"]] for c in a["components"])
        hsp = base * (1 + a["overhead_profit"])
        if abs(hsp - a["expected_hsp"]) > max(1.0, 0.005 * a["expected_hsp"]):
            bad.append((a["code"], round(hsp), a["expected_hsp"]))
    if bad:
        raise SystemExit(f"Verifikasi GAGAL untuk {len(bad)} analisa: {bad[:5]} — file tidak ditulis.")

    return {
        "_note": (
            "GOLDEN FIXTURE UJI dari RAB manual PLHUT Surakarta 2024 (ALFA.xlsx, sheet AHS). "
            "BUKAN data/template sistem (roadmap PAAX_ROADMAP_GAMBAR_KE_RAB Sec 0.1). "
            "Golden anchor: engine UMUM wajib reproduksi expected_hsp tiap analisa via (A+B+C)x(1+OP)."
        ),
        "source_file": "rab gedung plhut surakarta ALFA.xlsx :: sheet AHS",
        "region_hint": "surakarta-2024",
        "n_analyses": len(analyses),
        "resources": sorted(resources.values(), key=lambda x: x["code"]),
        "analyses": analyses,
    }


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    doc = build(sys.argv[1])
    OUT.write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK — {doc['n_analyses']} analisa terverifikasi, ditulis ke {OUT}")
