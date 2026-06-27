import io
from pathlib import Path

import openpyxl

from app.export.excel_exporter import export_rab_to_excel
from app.rab.loader import load_data
from app.rab.models import RABLineInput
from app.rab.rab import compute_hsp
from app.rab.sections import build_sectioned_rab

REPO_ROOT = Path(__file__).resolve().parents[3]
STORE = load_data(REPO_ROOT / "data")
BOOK = STORE.price_book("jateng")


def _row_by_value(ws, column: int, value: str) -> int:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=column).value == value:
            return row
    raise AssertionError(f"Value {value!r} not found in column {column} of {ws.title}")


def test_export_has_expected_sheets_and_formulas():
    lines = [
        RABLineInput(ahsp_code="AHSP.CK.001", volume=50, section="III"),
        RABLineInput(ahsp_code="AHSP.CK.002", volume=50, section="III"),
    ]
    result = build_sectioned_rab(
        lines, STORE.ahsp, BOOK, region="Jawa Tengah", region_code="jateng"
    )
    breakdowns = {
        code: compute_hsp(STORE.ahsp[code], BOOK)
        for code in ["AHSP.CK.001", "AHSP.CK.002"]
    }

    xlsx_bytes = export_rab_to_excel(result, breakdowns)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=False)

    assert wb.sheetnames[:4] == ["DKH", "HSP", "AHS", "HARGA BAHAN"]

    dkh = wb["DKH"]
    line_row = _row_by_value(dkh, 1, "Pasangan dinding bata merah 1/2 batu, camp. 1 PC : 5 PP")
    assert dkh.cell(line_row, 3).value == 50
    assert dkh.cell(line_row, 4).value.startswith("=HSP!F")
    assert dkh.cell(line_row, 5).value == 11.0
    assert dkh.cell(line_row, 6).value == f"=(C{line_row}*D{line_row})*E{line_row}/100"
    assert dkh.cell(line_row, 7).value == f"=(C{line_row}*D{line_row})+F{line_row}"

    subtotal_row = _row_by_value(dkh, 1, "Subtotal")
    assert dkh.cell(subtotal_row, 7).value.startswith("=SUMPRODUCT(")
    assert dkh.cell(subtotal_row + 1, 7).value.startswith("=SUM(")
    assert dkh.cell(subtotal_row + 2, 7).value.startswith("=SUM(")

    ahs = wb["AHS"]
    d_row = _row_by_value(ahs, 2, "D  Jumlah (A+B+C)")
    assert ahs.cell(d_row, 11).value.startswith("=K")
    f_row = _row_by_value(ahs, 2, "F  Harga Satuan Pekerjaan (D+E)")
    assert "ROUNDDOWN" in ahs.cell(f_row, 11).value

    harga = wb["HARGA BAHAN"]
    resource_rows = harga.max_row - 3
    expected_resource_count = len({
        component.resource_code
        for breakdown in breakdowns.values()
        for component in breakdown.components
    })
    assert resource_rows == expected_resource_count
    bata_row = _row_by_value(harga, 2, "BTA.01")
    assert harga.cell(bata_row, 6).value == 800
