"""
Template-based Excel export engine.

Uses openpyxl to generate formatted RAB spreadsheets.
"""

from __future__ import annotations

import io
from typing import Any

from app.domain.export.formatters import format_currency, format_number
from app.domain.rab.models import RABGroup, RABSummary


def generate_rab_excel(
    groups: list[RABGroup],
    summary: RABSummary,
    project_name: str = "Proyek",
    project_location: str = "Jakarta",
) -> bytes:
    """
    Generate a formatted RAB Excel workbook.

    Returns the workbook as bytes (for streaming download).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        # Fallback: return a minimal CSV-like bytes object
        return _generate_csv_fallback(groups, summary, project_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "RAB"

    # ── Styles ──────────────────────────────────────────────
    header_font = Font(name="Arial", bold=True, size=14)
    subheader_font = Font(name="Arial", bold=True, size=11)
    bold_font = Font(name="Arial", bold=True, size=10)
    normal_font = Font(name="Arial", size=10)
    currency_fmt = '#,##0'

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font_white = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    group_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    summary_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ── Column widths ───────────────────────────────────────
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 22

    # ── Title section ───────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = "RENCANA ANGGARAN BIAYA (RAB)"
    ws["A1"].font = header_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Proyek: {project_name} — Lokasi: {project_location}"
    ws["A2"].font = subheader_font
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Table header ────────────────────────────────────────
    row = 4
    headers = ["No", "Uraian Pekerjaan", "Satuan", "Volume", "Harga Satuan (Rp)", "Jumlah (Rp)"]
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header_text)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row += 1

    # ── Data rows ───────────────────────────────────────────
    for group in groups:
        # Group header row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = ws.cell(row=row, column=1, value=group.nama)
        cell.font = bold_font
        cell.fill = group_fill
        cell.border = thin_border
        for c in range(2, 7):
            ws.cell(row=row, column=c).fill = group_fill
            ws.cell(row=row, column=c).border = thin_border
        row += 1

        for item in group.items:
            item.hitung_jumlah()
            ws.cell(row=row, column=1, value=item.kode).font = normal_font
            ws.cell(row=row, column=2, value=item.uraian).font = normal_font
            ws.cell(row=row, column=3, value=item.satuan.value).font = normal_font
            ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

            vol_cell = ws.cell(row=row, column=4, value=item.volume)
            vol_cell.font = normal_font
            vol_cell.number_format = '#,##0.00'

            price_cell = ws.cell(row=row, column=5, value=item.harga_satuan)
            price_cell.font = normal_font
            price_cell.number_format = currency_fmt

            amount_cell = ws.cell(row=row, column=6, value=item.jumlah)
            amount_cell.font = normal_font
            amount_cell.number_format = currency_fmt

            for c in range(1, 7):
                ws.cell(row=row, column=c).border = thin_border
            row += 1

        # Group subtotal
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        st_cell = ws.cell(row=row, column=1, value=f"Subtotal {group.nama}")
        st_cell.font = bold_font
        st_cell.alignment = Alignment(horizontal="right")
        st_amount = ws.cell(row=row, column=6, value=group.subtotal)
        st_amount.font = bold_font
        st_amount.number_format = currency_fmt
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = thin_border
        row += 1
        row += 1  # blank row

    # ── Summary section ─────────────────────────────────────
    summary_rows = [
        ("SUBTOTAL", summary.subtotal),
        (f"PPN ({summary.ppn_rate:.0%})", summary.ppn),
        (f"Kontingensi ({summary.contingency_rate:.0%})", summary.contingency),
        (f"Overhead & Profit ({summary.overhead_profit_rate:.0%})", summary.overhead_profit),
        ("GRAND TOTAL", summary.grand_total),
    ]

    for label, value in summary_rows:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="right")
        cell.fill = summary_fill

        val_cell = ws.cell(row=row, column=6, value=value)
        val_cell.font = bold_font
        val_cell.number_format = currency_fmt
        val_cell.fill = summary_fill

        for c in range(1, 7):
            ws.cell(row=row, column=c).border = thin_border
        row += 1

    # ── Save to bytes ───────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _generate_csv_fallback(
    groups: list[RABGroup],
    summary: RABSummary,
    project_name: str,
) -> bytes:
    """Fallback CSV generation when openpyxl is not available."""
    lines = [
        f"RENCANA ANGGARAN BIAYA (RAB) — {project_name}",
        "Kode,Uraian,Satuan,Volume,Harga Satuan,Jumlah",
    ]
    for group in groups:
        lines.append(f"\n{group.nama}")
        for item in group.items:
            item.hitung_jumlah()
            lines.append(
                f"{item.kode},{item.uraian},{item.satuan.value},"
                f"{item.volume},{format_currency(item.harga_satuan)},{format_currency(item.jumlah)}"
            )
        lines.append(f"Subtotal,,,,, {format_currency(group.subtotal)}")

    lines.append(f"\nSUBTOTAL,,,,, {format_currency(summary.subtotal)}")
    lines.append(f"PPN,,,,, {format_currency(summary.ppn)}")
    lines.append(f"Kontingensi,,,,, {format_currency(summary.contingency)}")
    lines.append(f"GRAND TOTAL,,,,, {format_currency(summary.grand_total)}")

    return "\n".join(lines).encode("utf-8")
