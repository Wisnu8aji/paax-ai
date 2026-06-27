from __future__ import annotations

from io import BytesIO
from typing import Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..rab.models import ComponentCost, HSPBreakdown, RABLine
from ..rab.sections import SectionedRABResult


CategoryBlock = Tuple[str, str]
CATEGORY_BLOCKS: List[CategoryBlock] = [
    ("upah", "A. Upah Tenaga"),
    ("bahan", "B. Bahan"),
    ("alat", "C. Peralatan"),
]


def _ordered_lines(result: SectionedRABResult) -> List[RABLine]:
    return [line for section in result.sections for line in section.lines]


def _unique_codes(lines: Iterable[RABLine]) -> List[str]:
    seen = set()
    codes: List[str] = []
    for line in lines:
        if line.ahsp_code not in seen:
            seen.add(line.ahsp_code)
            codes.append(line.ahsp_code)
    return codes


def _unique_resources(
    codes: Iterable[str],
    breakdowns: Dict[str, HSPBreakdown],
) -> List[ComponentCost]:
    seen = set()
    resources: List[ComponentCost] = []
    for code in codes:
        breakdown = breakdowns[code]
        for component in breakdown.components:
            if component.resource_code in seen:
                continue
            seen.add(component.resource_code)
            resources.append(component)
    return resources


def _write_headers(ws: Worksheet, row: int, values: Dict[int, object]) -> None:
    for column, value in values.items():
        ws.cell(row=row, column=column, value=value)


def _write_harga_bahan(ws: Worksheet, resources: List[ComponentCost]) -> None:
    ws["A1"] = "DAFTAR HARGA BAHAN DAN UPAH"
    _write_headers(ws, 3, {
        2: "Kode",
        3: "Nama",
        4: "Kategori",
        5: "Satuan",
        6: "Harga (Rp)",
    })
    for offset, resource in enumerate(resources, start=4):
        ws.cell(offset, 2, resource.resource_code)
        ws.cell(offset, 3, resource.resource_name)
        ws.cell(offset, 4, resource.category)
        ws.cell(offset, 5, resource.unit)
        ws.cell(offset, 6, resource.unit_price)


def _write_component_rows(
    ws: Worksheet,
    start_row: int,
    components: List[ComponentCost],
) -> Tuple[int, int | None, int | None]:
    row = start_row
    first = None
    last = None
    for idx, component in enumerate(components, start=1):
        if first is None:
            first = row
        last = row
        ws.cell(row, 3, f"{idx}.")
        ws.cell(row, 4, component.resource_name)
        ws.cell(row, 7, component.resource_code)
        ws.cell(row, 8, component.unit)
        ws.cell(row, 9, component.coefficient)
        ws.cell(row, 10, f"=VLOOKUP(D{row},'HARGA BAHAN'!$C$4:$F$9999,4,0)")
        ws.cell(row, 11, f"=I{row}*J{row}")
        row += 1
    return row, first, last


def _write_category_block(
    ws: Worksheet,
    row: int,
    label: str,
    components: List[ComponentCost],
    summary_label: str,
) -> Tuple[int, int]:
    ws.cell(row, 2, label)
    row += 1
    row, first, last = _write_component_rows(ws, row, components)
    summary_row = row
    ws.cell(summary_row, 2, summary_label)
    if first is None or last is None:
        ws.cell(summary_row, 11, 0)
    else:
        ws.cell(summary_row, 11, f"=SUM(K{first}:K{last})")
    return summary_row, row + 1


def _write_ahs(
    ws: Worksheet,
    codes: List[str],
    breakdowns: Dict[str, HSPBreakdown],
) -> Dict[str, int]:
    row = 1
    f_rows: Dict[str, int] = {}
    for code in codes:
        breakdown = breakdowns[code]
        ws.cell(row, 2, " Jenis Pekerjaan")
        ws.cell(row, 5, ":")
        ws.cell(row, 6, breakdown.name)
        ws.cell(row + 1, 2, "Satuan Pekerjaan")
        ws.cell(row + 1, 5, ":")
        ws.cell(row + 1, 6, breakdown.unit)
        _write_headers(ws, row + 3, {
            2: "Uraian",
            7: "Kode",
            8: "Satuan",
            9: "Koefisien",
            10: "Harga (Rp)",
            11: "Jumlah (Rp)",
        })

        cursor = row + 4
        sum_rows: Dict[str, int] = {}
        for category, label in CATEGORY_BLOCKS:
            components = [c for c in breakdown.components if c.category == category]
            summary_label = {
                "upah": "Jumlah Harga Upah Tenaga",
                "bahan": "Jumlah Harga Bahan",
                "alat": "Jumlah Harga Peralatan",
            }[category]
            summary_row, cursor = _write_category_block(
                ws, cursor, label, components, summary_label
            )
            sum_rows[category] = summary_row

        d_row = cursor
        ws.cell(d_row, 2, "D  Jumlah (A+B+C)")
        ws.cell(d_row, 11, f"=K{sum_rows['upah']}+K{sum_rows['bahan']}+K{sum_rows['alat']}")

        e_row = d_row + 1
        ws.cell(e_row, 2, "E  Overhead & Profit")
        ws.cell(e_row, 9, breakdown.overhead_profit)
        ws.cell(e_row, 10, "x D")
        ws.cell(e_row, 11, f"=I{e_row}*K{d_row}")

        f_row = e_row + 1
        ws.cell(f_row, 2, "F  Harga Satuan Pekerjaan (D+E)")
        ws.cell(f_row, 11, f"=ROUNDDOWN((K{d_row}+K{e_row}),0)")
        f_rows[code] = f_row
        row = f_row + 3
    return f_rows


def _write_hsp(
    ws: Worksheet,
    codes: List[str],
    breakdowns: Dict[str, HSPBreakdown],
    ahs_f_rows: Dict[str, int],
) -> Dict[str, int]:
    ws["A1"] = "HARGA SATUAN PEKERJAAN"
    _write_headers(ws, 3, {
        1: "NO",
        2: "URAIAN PEKERJAAN",
        5: "SAT.",
        6: "HARGA SATUAN (Rp.)",
    })
    hsp_rows: Dict[str, int] = {}
    for idx, code in enumerate(codes, start=1):
        row = idx + 3
        breakdown = breakdowns[code]
        ws.cell(row, 1, idx)
        ws.cell(row, 2, breakdown.name)
        ws.cell(row, 5, breakdown.unit)
        ws.cell(row, 6, f"=AHS!K{ahs_f_rows[code]}")
        hsp_rows[code] = row
    return hsp_rows


def _write_dkh(
    ws: Worksheet,
    result: SectionedRABResult,
    hsp_rows: Dict[str, int],
) -> None:
    ws["A1"] = "DAFTAR KUANTITAS DAN HARGA"
    _write_headers(ws, 3, {
        1: "Jenis barang/jasa",
        2: "Satuan",
        3: "Volume",
        4: "Harga satuan (Rp.)",
        5: "Pajak (%)",
        6: "Pajak (Rp.)",
        7: "Total (Rp.)",
    })

    row = 4
    first_data_row = None
    last_data_row = None
    tax_pct = result.ppn_rate * 100
    for section in result.sections:
        ws.cell(row, 1, section.title)
        for column in range(3, 8):
            ws.cell(row, column, 0)
        row += 1
        for line in section.lines:
            if first_data_row is None:
                first_data_row = row
            last_data_row = row
            ws.cell(row, 1, line.name)
            ws.cell(row, 2, line.unit)
            ws.cell(row, 3, line.volume)
            ws.cell(row, 4, f"=HSP!F{hsp_rows[line.ahsp_code]}")
            ws.cell(row, 5, tax_pct)
            ws.cell(row, 6, f"=(C{row}*D{row})*E{row}/100")
            ws.cell(row, 7, f"=(C{row}*D{row})+F{row}")
            row += 1

    if first_data_row is None or last_data_row is None:
        first_data_row = last_data_row = row
        ws.cell(row, 3, 0)
        ws.cell(row, 4, 0)
        ws.cell(row, 6, 0)
        ws.cell(row, 7, 0)

    ws.cell(row, 1, "Subtotal")
    ws.cell(row, 7, f"=SUMPRODUCT(C{first_data_row}:C{last_data_row},D{first_data_row}:D{last_data_row})")
    ws.cell(row + 1, 1, "PPN")
    ws.cell(row + 1, 7, f"=SUM(F{first_data_row}:F{last_data_row})")
    ws.cell(row + 2, 1, "Total")
    ws.cell(row + 2, 7, f"=SUM(G{first_data_row}:G{last_data_row})")


def _apply_basic_widths(workbook: Workbook) -> None:
    widths = {
        "A": 34,
        "B": 16,
        "C": 14,
        "D": 42,
        "E": 14,
        "F": 18,
        "G": 18,
        "H": 12,
        "I": 12,
        "J": 18,
        "K": 18,
    }
    for ws in workbook.worksheets:
        for column, width in widths.items():
            ws.column_dimensions[column].width = width


def export_rab_to_excel(
    result: SectionedRABResult,
    breakdowns: Dict[str, HSPBreakdown],
) -> bytes:
    """Bangun workbook .xlsx dengan sheet HARGA BAHAN, AHS, HSP, dan DKH."""
    lines = _ordered_lines(result)
    codes = _unique_codes(lines)
    for code in codes:
        if code not in breakdowns:
            raise KeyError(f"Breakdown HSP untuk '{code}' tidak tersedia.")

    workbook = Workbook()
    dkh = workbook.active
    dkh.title = "DKH"
    hsp = workbook.create_sheet("HSP")
    ahs = workbook.create_sheet("AHS")
    harga_bahan = workbook.create_sheet("HARGA BAHAN")

    resources = _unique_resources(codes, breakdowns)
    _write_harga_bahan(harga_bahan, resources)
    ahs_f_rows = _write_ahs(ahs, codes, breakdowns)
    hsp_rows = _write_hsp(hsp, codes, breakdowns, ahs_f_rows)
    _write_dkh(dkh, result, hsp_rows)
    _apply_basic_widths(workbook)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
